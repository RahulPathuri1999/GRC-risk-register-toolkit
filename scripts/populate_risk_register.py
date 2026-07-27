"""
Append new risks to a risk register built from the organizational template
(33-column schema: Asset Evaluation | Risk Identification | Risk Analysis |
Risk Mitigation | Residual Risk Assessment | Residual Risk Description).

Key differences from the simpler schema this project started with:
  - Risk = Asset Value (C+I+A) x Likelihood x Consequence (3-factor, not 2-factor)
  - "Threats" and "Vulnerabilities" are separate columns, not one description field
  - Asset ID/Name/Owner/Date are only filled on the FIRST row for a given asset;
    subsequent risk rows for that asset leave those columns blank (but C/I/A/Asset
    Value ARE repeated on every row for that asset, per the template's own example data)
  - Risk Value/Risk Level and Post-Mit Risk Value/Risk Level had NO formulas in the
    template as uploaded -- this script adds them as real Excel formulas so the sheet
    recalculates itself, instead of requiring the numbers to be typed in by hand

Usage:
    python populate_risk_register.py <register.xlsx> <risks.json> [--apply]

Without --apply (default): prints a JSON report of pre-flight checks and writes nothing.
With --apply: writes the rows after printing the same report.

risks.json format (a list of objects):
[
  {
    "asset_id": "A-002",                 # required. Reuse an existing ID to add another
                                          # risk to that asset; use a new ID to add an asset.
    "asset_name": "Finance Shared Drive", # required only when asset_id is new
    "asset_owner": "Priya",               # required only when asset_id is new
    "identification_date": "2026-07-27",  # optional, defaults to today; only used when new
    "confidentiality": 3, "integrity": 2, "availability": 1,   # 1-3 each; required when new
    "risk_id": "R-006",                   # optional, auto-numbered if omitted
    "threat": "Unauthorized access to shared financial files",
    "vulnerability": "No folder-level permission restrictions",
    "likelihood": 3, "consequence": 4,    # 1-5 each
    "risk_treatment_option": "Mitigate",  # Accept / Avoid / Transfer / Mitigate
    "existing_controls": "None",
    "mitigation_plan": "Restrict folder access to Finance team via AD group",
    "iso_control": "A.9.1.2",
    "mitigation_responsibility": "IT Ops",
    "mitigation_target_date": "2026-09-01",
    "mitigation_status": "Pending Approval",
    "risk_status": "Open",
    "residual_likelihood": 1, "residual_impact": 3,
    "residual_description": "Access limited to Finance team only",
    "residual_control": "AD group + quarterly access review",
    "risk_control_responsibility": "IT Ops",
    "control_target_date": "2026-09-15",
    "control_status": "Pending Approval",
    "residual_risk_status": "Monitored"
  }
]
Only asset_id + threat are truly required; everything else is optional and left blank if
omitted. Risk Value / Risk Level / Post-Mit Risk Value / Risk Level are always written as
formulas, never computed in Python.
"""
import sys
import json
import re
import argparse
import datetime
from difflib import SequenceMatcher
import openpyxl

DATA_START_ROW = 8  # rows 3-7 in this template are the reference/legend rows, not data

COLS = {
    "asset_id": 1, "identification_date": 2, "asset_name": 3, "asset_owner": 4,
    "confidentiality": 5, "integrity": 6, "availability": 7, "asset_value": 8,
    "risk_id": 9, "threat": 10, "vulnerability": 11,
    "likelihood": 12, "consequence": 13, "risk_value": 14, "risk_level": 15,
    "risk_treatment_option": 16, "existing_controls": 17, "mitigation_plan": 18,
    "iso_control": 19, "mitigation_responsibility": 20, "mitigation_target_date": 21,
    "mitigation_status": 22, "risk_status": 23,
    "residual_likelihood": 24, "residual_impact": 25, "post_mit_risk_value": 26,
    "residual_risk_level": 27, "residual_description": 28, "residual_control": 29,
    "risk_control_responsibility": 30, "control_target_date": 31, "control_status": 32,
    "residual_risk_status": 33,
}
COL_LETTER = {k: openpyxl.utils.get_column_letter(v) for k, v in COLS.items()}

STOPWORDS = {
    "the", "a", "an", "to", "of", "in", "on", "could", "would", "and", "or", "due",
    "via", "for", "is", "are", "be", "by", "with", "into", "leading", "that", "used",
}

CONTROL_KEYWORDS = {
    "mfa": {"password", "credential", "login", "authentication", "access"},
    "multi-factor": {"password", "credential", "login", "authentication", "access"},
    "encryption": {"data", "database", "storage", "laptop", "device", "media"},
    "training": {"phishing", "social", "employee", "staff", "human", "error"},
    "awareness": {"phishing", "social", "employee", "staff", "human", "error"},
    "backup": {"availability", "outage", "disaster", "downtime", "power"},
    "antivirus": {"malware", "virus", "attack"},
    "patch": {"malware", "vulnerability", "configuration", "attack"},
    "access": {"theft", "unauthorized", "physical", "access"},
    "group": {"theft", "unauthorized", "access", "permission"},
}


def tokenize(text):
    words = re.findall(r"[a-z]+", (text or "").lower())
    return {w for w in words if w not in STOPWORDS and len(w) > 2}


def text_similarity(a, b):
    ta, tb = tokenize(a), tokenize(b)
    jaccard = len(ta & tb) / len(ta | tb) if (ta or tb) else 0
    seq = SequenceMatcher(None, a or "", b or "").ratio()
    return round(0.7 * jaccard + 0.3 * seq, 2)


def first_empty_row(ws):
    row = DATA_START_ROW
    while ws.cell(row=row, column=COLS["risk_id"]).value not in (None, ""):
        row += 1
    return row


def load_existing_rows(ws, up_to_row):
    """Forward-fills Asset ID/Name/Owner/CIA/Asset Value, since the template only fills
    those on the first row of each asset group and leaves them blank on the rest."""
    rows = []
    last = {}
    for r in range(DATA_START_ROW, up_to_row):
        if ws.cell(row=r, column=COLS["risk_id"]).value in (None, ""):
            continue
        row = {k: ws.cell(row=r, column=c).value for k, c in COLS.items()}
        for field in ("asset_id", "asset_name", "asset_owner", "confidentiality",
                      "integrity", "availability", "asset_value"):
            if row[field] in (None, ""):
                row[field] = last.get(field)
            else:
                last[field] = row[field]
        row["_row"] = r
        rows.append(row)
    return rows


def duplicate_check(new_risk, existing_rows, threshold=0.35, same_treatment_option_threshold=0.15):
    """Combines Threat + Vulnerability into one text for comparison, since together they
    describe the same underlying risk that a single 'description' field would in a simpler
    schema."""
    new_text = f"{new_risk.get('threat', '')} {new_risk.get('vulnerability', '')}"
    flags = []
    for row in existing_rows:
        if row["asset_id"] != new_risk.get("asset_id"):
            continue
        existing_text = f"{row.get('threat', '')} {row.get('vulnerability', '')}"
        sim = text_similarity(new_text, existing_text)
        same_treatment = row.get("risk_treatment_option") == new_risk.get("risk_treatment_option")
        bar = same_treatment_option_threshold if same_treatment else threshold
        if sim >= bar:
            flags.append({
                "risk_id": row["risk_id"], "threat": row.get("threat"),
                "vulnerability": row.get("vulnerability"), "similarity": sim,
            })
    flags.sort(key=lambda f: -f["similarity"])
    return flags


def control_propagation_check(new_risk, existing_rows):
    plan_text = (new_risk.get("mitigation_plan") or "").lower()
    matched = [kw for kw in CONTROL_KEYWORDS if kw in plan_text]
    if not matched:
        return []
    related_terms = set().union(*(CONTROL_KEYWORDS[kw] for kw in matched))
    flags = []
    for row in existing_rows:
        if row["asset_id"] != new_risk.get("asset_id") or row["risk_id"] == new_risk.get("risk_id"):
            continue
        candidate_text = f"{row.get('threat', '')} {row.get('vulnerability', '')}"
        if tokenize(candidate_text) & related_terms:
            flags.append({
                "risk_id": row["risk_id"], "threat": row.get("threat"),
                "current_residual_risk_level_formula_row": row["_row"],
                "reason": f"Mitigation plan mentions {matched}, which plausibly also "
                          f"mitigates this risk's cause",
            })
    return flags


GENERIC_ASSET_WORDS = {
    "company", "corporate", "employee", "employees", "system", "systems", "data",
    "service", "services", "management", "platform", "process", "application", "applications",
}


def blast_radius_check(new_risk, existing_rows):
    is_new_asset = not any(r["asset_id"] == new_risk.get("asset_id") for r in existing_rows)
    if not is_new_asset:
        return None
    new_tokens = tokenize(new_risk.get("asset_name", "")) - GENERIC_ASSET_WORDS
    related, seen = [], set()
    for row in existing_rows:
        if row["asset_id"] in seen:
            continue
        candidate = (tokenize(row.get("asset_name", "")) |
                     tokenize(row.get("threat", "")) |
                     tokenize(row.get("vulnerability", ""))) - GENERIC_ASSET_WORDS
        if new_tokens & candidate:
            related.append(row["asset_id"])
            seen.add(row["asset_id"])
    return {"new_asset": True, "possibly_related_existing_assets": related}


def run_checks(new_risk, existing_rows):
    return {
        "risk_id": new_risk.get("risk_id"),
        "duplicate_check": duplicate_check(new_risk, existing_rows),
        "control_propagation_check": control_propagation_check(new_risk, existing_rows),
        "blast_radius_check": blast_radius_check(new_risk, existing_rows),
    }


def write_row(ws, row, risk, fallback_id, is_new_asset):
    r = row
    if is_new_asset:
        ws.cell(row=r, column=COLS["asset_id"], value=risk.get("asset_id"))
        ws.cell(row=r, column=COLS["identification_date"],
                 value=risk.get("identification_date", datetime.date.today().isoformat()))
        ws.cell(row=r, column=COLS["asset_name"], value=risk.get("asset_name", ""))
        ws.cell(row=r, column=COLS["asset_owner"], value=risk.get("asset_owner", ""))
    # C/I/A/Asset Value repeat on every row for the asset, per the template's own convention
    ws.cell(row=r, column=COLS["confidentiality"], value=risk.get("confidentiality"))
    ws.cell(row=r, column=COLS["integrity"], value=risk.get("integrity"))
    ws.cell(row=r, column=COLS["availability"], value=risk.get("availability"))
    ws.cell(row=r, column=COLS["asset_value"],
             value=f"={COL_LETTER['confidentiality']}{r}+{COL_LETTER['integrity']}{r}+{COL_LETTER['availability']}{r}")

    ws.cell(row=r, column=COLS["risk_id"], value=risk.get("risk_id", fallback_id))
    ws.cell(row=r, column=COLS["threat"], value=risk.get("threat", ""))
    ws.cell(row=r, column=COLS["vulnerability"], value=risk.get("vulnerability", ""))
    ws.cell(row=r, column=COLS["likelihood"], value=risk.get("likelihood"))
    ws.cell(row=r, column=COLS["consequence"], value=risk.get("consequence"))
    ws.cell(row=r, column=COLS["risk_value"],
             value=f"={COL_LETTER['asset_value']}{r}*{COL_LETTER['likelihood']}{r}*{COL_LETTER['consequence']}{r}")
    ws.cell(row=r, column=COLS["risk_level"], value=(
        f'=IF({COL_LETTER["risk_value"]}{r}>=225,"Extreme",'
        f'IF({COL_LETTER["risk_value"]}{r}>=128,"Major",'
        f'IF({COL_LETTER["risk_value"]}{r}>=37,"Medium","Minor")))'
    ))
    ws.cell(row=r, column=COLS["risk_treatment_option"], value=risk.get("risk_treatment_option", ""))
    ws.cell(row=r, column=COLS["existing_controls"], value=risk.get("existing_controls", ""))
    ws.cell(row=r, column=COLS["mitigation_plan"], value=risk.get("mitigation_plan", ""))
    ws.cell(row=r, column=COLS["iso_control"], value=risk.get("iso_control", ""))
    ws.cell(row=r, column=COLS["mitigation_responsibility"], value=risk.get("mitigation_responsibility", ""))
    ws.cell(row=r, column=COLS["mitigation_target_date"], value=risk.get("mitigation_target_date", ""))
    ws.cell(row=r, column=COLS["mitigation_status"], value=risk.get("mitigation_status", ""))
    ws.cell(row=r, column=COLS["risk_status"], value=risk.get("risk_status", "Open"))

    ws.cell(row=r, column=COLS["residual_likelihood"], value=risk.get("residual_likelihood"))
    ws.cell(row=r, column=COLS["residual_impact"], value=risk.get("residual_impact"))
    ws.cell(row=r, column=COLS["post_mit_risk_value"],
             value=f"={COL_LETTER['asset_value']}{r}*{COL_LETTER['residual_likelihood']}{r}*{COL_LETTER['residual_impact']}{r}")
    ws.cell(row=r, column=COLS["residual_risk_level"], value=(
        f'=IF({COL_LETTER["post_mit_risk_value"]}{r}>=225,"Extreme",'
        f'IF({COL_LETTER["post_mit_risk_value"]}{r}>=128,"Major",'
        f'IF({COL_LETTER["post_mit_risk_value"]}{r}>=37,"Medium","Minor")))'
    ))
    ws.cell(row=r, column=COLS["residual_description"], value=risk.get("residual_description", ""))
    ws.cell(row=r, column=COLS["residual_control"], value=risk.get("residual_control", ""))
    ws.cell(row=r, column=COLS["risk_control_responsibility"], value=risk.get("risk_control_responsibility", ""))
    ws.cell(row=r, column=COLS["control_target_date"], value=risk.get("control_target_date", ""))
    ws.cell(row=r, column=COLS["control_status"], value=risk.get("control_status", ""))
    ws.cell(row=r, column=COLS["residual_risk_status"], value=risk.get("residual_risk_status", ""))


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx_path")
    ap.add_argument("json_path")
    ap.add_argument("--apply", action="store_true")
    args = ap.parse_args()

    with open(args.json_path) as f:
        risks = json.load(f)
    if isinstance(risks, dict):
        risks = [risks]

    wb = openpyxl.load_workbook(args.xlsx_path)
    ws = wb["Risk Register"]

    start_row = first_empty_row(ws)
    existing_rows = load_existing_rows(ws, start_row)

    reports = []
    row = start_row
    for risk in risks:
        report = run_checks(risk, existing_rows)
        reports.append(report)
        fallback_id = f"R-{row - DATA_START_ROW + 1:03d}"
        is_new_asset = not any(r["asset_id"] == risk.get("asset_id") for r in existing_rows)
        if args.apply:
            write_row(ws, row, risk, fallback_id, is_new_asset)
        existing_rows.append({
            **{k: risk.get(k) for k in COLS},
            "risk_id": risk.get("risk_id", fallback_id),
            "_row": row,
        })
        if args.apply:
            row += 1

    print(json.dumps(reports, indent=2, default=str))

    if args.apply:
        wb.save(args.xlsx_path)
        print(f"\nAppended {len(risks)} risk(s) to {args.xlsx_path}. Remember to run recalc.py.", file=sys.stderr)


if __name__ == "__main__":
    main()
